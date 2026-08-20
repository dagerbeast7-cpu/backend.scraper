"""Test workbook cleanup logic on live downloaded workbook bytes."""
import io
import sys
sys.path.insert(0, ".")
import httpx
import phonenumbers
from openpyxl import load_workbook

from app.dedup.engine import normalize_phone

# Fetch current live workbook bytes
res = httpx.get("https://backendscraper-production.up.railway.app/prospects/export/download", timeout=30.0)
assert res.status_code == 200, f"Failed to download: {res.status_code}"

wb = load_workbook(io.BytesIO(res.content))
ws = wb["Prospects"]

rows_before = ws.max_row - 1
print(f"Rows before cleanup: {rows_before}")

removed_count = 0
removed_leads = []

# Iterate from bottom to top to preserve row indices
for row_idx in range(ws.max_row, 1, -1):
    raw_phone = ws.cell(row=row_idx, column=3).value  # Column C is Phone
    bus_name = ws.cell(row=row_idx, column=1).value  # Column A is Business Name
    
    p_norm = normalize_phone(str(raw_phone)) if raw_phone else None
    if not p_norm:
        removed_count += 1
        removed_leads.append((row_idx, bus_name, raw_phone))
        ws.delete_rows(row_idx, 1)

rows_after = ws.max_row - 1
print(f"Rows removed: {removed_count}")
print(f"Rows remaining: {rows_after}")

# Verify remaining rows all have valid phone
invalid_remaining = 0
for row_idx in range(2, ws.max_row + 1):
    p_val = ws.cell(row=row_idx, column=3).value
    if not p_val or not normalize_phone(str(p_val)):
        invalid_remaining += 1

print(f"Invalid remaining: {invalid_remaining}")
assert invalid_remaining == 0, "All remaining rows must have usable phone numbers"
assert rows_after == rows_before - removed_count, "Row math must match"

print("CLEANUP LOGIC VERIFIED SUCCESSFULLY!")
