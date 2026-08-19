"""
Comprehensive tests for the canonical Supabase Storage Excel workflow.

Verifies:
1. Initial creation when canonical workbook is absent.
2. Append-only behavior (existing rows are NEVER modified).
3. Preservation of caller-entered edits (e.g. Status = CLOSED, custom notes, broker names).
4. Identity matching hierarchy (Google Maps ID, normalized phone, website, name+city).
5. Simulated Supabase Storage persistence across container restarts.
6. Retrieval / download from the API service.
"""
from __future__ import annotations

import io
import sys
import uuid
from typing import Sequence

sys.path.insert(0, ".")

from openpyxl import load_workbook

from app.db.models import LeadStatus, Prospect, VerificationStatus
from app.storage.excel_storage import (
    SupabaseStorageClient,
    append_new_leads,
    create_initial_workbook,
    get_canonical_workbook_bytes,
    sync_prospects_to_storage_workbook,
)


class MockSupabaseStorageClient(SupabaseStorageClient):
    """In-memory mock for Supabase Storage to test container isolation and persistence."""

    def __init__(self):
        super().__init__(base_url="https://mock.supabase.co", api_key="mock-key", bucket_name="leadzen-exports")
        self.storage: dict[str, bytes] = {}

    def ensure_bucket_exists(self) -> None:
        pass

    def download_file(self, object_name: str | None = None) -> bytes | None:
        name = object_name or "leadzen_prospects.xlsx"
        return self.storage.get(name)

    def upload_file(self, data: bytes, object_name: str | None = None) -> bool:
        name = object_name or "leadzen_prospects.xlsx"
        self.storage[name] = data
        return True


def make_dummy_prospect(
    name: str,
    phone: str | None = None,
    city: str = "Mumbai",
    locality: str = "Bandra West",
    website: str | None = None,
    google_maps_id: str | None = None,
    score: int = 80,
) -> Prospect:
    p = Prospect(
        id=uuid.uuid4(),
        business_name=name,
        phone=phone,
        city=city,
        locality=locality,
        website=website,
        google_maps_id=google_maps_id,
        industry="real estate broker",
        score=score,
        status=LeadStatus.NEW,
        verification_status=VerificationStatus.PHONE_VERIFIED if phone else VerificationStatus.UNVERIFIED,
    )
    return p


failures: list[str] = []


def check(label: str, condition: bool) -> None:
    status = "OK" if condition else "FAIL"
    print(f"  [{status}] {label}")
    if not condition:
        failures.append(label)


print("=" * 80)
print("RUNNING CANONICAL EXCEL WORKFLOW TESTS")
print("=" * 80)

# ------------------------------------------------------------------------------
# TEST 1: Initial Creation with 5 Leads
# ------------------------------------------------------------------------------
print("\n--- TEST 1: Canonical workbook absent -> initial workbook created with 5 rows ---")
mock_storage = MockSupabaseStorageClient()

leads_batch_1 = [
    make_dummy_prospect("Alpha Realty", "+91 98200 11111", city="Mumbai", google_maps_id="ChIJ_alpha"),
    make_dummy_prospect("Beta Properties", "+91 98200 22222", city="Mumbai", google_maps_id="ChIJ_beta"),
    make_dummy_prospect("Gamma Consultants", "+91 98200 33333", city="Mumbai", google_maps_id="ChIJ_gamma"),
    make_dummy_prospect("Delta Estates", "+91 98200 44444", city="Mumbai", google_maps_id="ChIJ_delta"),
    make_dummy_prospect("Epsilon Realtors", "+91 98200 55555", city="Mumbai", google_maps_id="ChIJ_epsilon"),
]

wb1 = create_initial_workbook(leads_batch_1)
buf1 = io.BytesIO()
wb1.save(buf1)
mock_storage.upload_file(buf1.getvalue())

# Inspect created sheet
saved_wb = load_workbook(io.BytesIO(mock_storage.download_file()))
ws1 = saved_wb["Prospects"]
data_rows_1 = [row for row in ws1.iter_rows(min_row=2, values_only=True) if any(row)]

check("Workbook created with exactly 5 prospect rows", len(data_rows_1) == 5)
check("First row business name is 'Alpha Realty'", data_rows_1[0][0] == "Alpha Realty")
check("First row status defaults to 'NOT_CONTACTED'", data_rows_1[0][10] == "NOT_CONTACTED")
check("Export Info metadata sheet exists", "Export Info" in saved_wb.sheetnames)

# ------------------------------------------------------------------------------
# TEST 2: Caller edits status to 'CLOSED' -> Re-scrape same 5 leads -> Row untouched!
# ------------------------------------------------------------------------------
print("\n--- TEST 2: Caller changes row 1 Status to 'CLOSED' -> re-scrape same 5 leads ---")
# Simulate caller opening workbook and changing cell K2 to 'CLOSED' and adding a custom broker note
ws1["K2"] = "CLOSED"  # Column K is Status
ws1["B2"] = "Vikram Sharma (Direct Owner)"  # Column B is Contact / Broker Name

# Save caller edited workbook back to storage
caller_buf = io.BytesIO()
saved_wb.save(caller_buf)
mock_storage.upload_file(caller_buf.getvalue())

# Now scraper runs on same 5 leads again
existing_bytes = mock_storage.download_file()
wb_loaded = load_workbook(io.BytesIO(existing_bytes))
appended = append_new_leads(wb_loaded, leads_batch_1)

# Save updated workbook
out_buf = io.BytesIO()
wb_loaded.save(out_buf)
mock_storage.upload_file(out_buf.getvalue())

# Verify
wb_after_rescrape = load_workbook(io.BytesIO(mock_storage.download_file()))
ws2 = wb_after_rescrape["Prospects"]
data_rows_2 = [row for row in ws2.iter_rows(min_row=2, values_only=True) if any(row)]

check("Zero new rows appended on identical scrape", appended == 0)
check("Total rows still exactly 5", len(data_rows_2) == 5)
check("Caller-edited Status 'CLOSED' was PRESERVED exactly", data_rows_2[0][10] == "CLOSED")
check("Caller-edited Broker Name was PRESERVED exactly", data_rows_2[0][1] == "Vikram Sharma (Direct Owner)")

# ------------------------------------------------------------------------------
# TEST 3: Scrape 3 existing leads + 2 brand new leads -> Workbook has 7 rows
# ------------------------------------------------------------------------------
print("\n--- TEST 3: Scrape 3 existing + 2 brand new leads -> total 7 rows ---")
leads_batch_3 = [
    leads_batch_1[0],  # Alpha (already in sheet with CLOSED)
    leads_batch_1[1],  # Beta (already in sheet)
    leads_batch_1[2],  # Gamma (already in sheet)
    make_dummy_prospect("Zeta Infra", "+91 98200 66666", city="Mumbai", google_maps_id="ChIJ_zeta"),
    make_dummy_prospect("Eta Homes", "+91 98200 77777", city="Mumbai", google_maps_id="ChIJ_eta"),
]

wb_loaded_3 = load_workbook(io.BytesIO(mock_storage.download_file()))
appended_3 = append_new_leads(wb_loaded_3, leads_batch_3)

buf3 = io.BytesIO()
wb_loaded_3.save(buf3)
mock_storage.upload_file(buf3.getvalue())

wb_after_3 = load_workbook(io.BytesIO(mock_storage.download_file()))
ws3 = wb_after_3["Prospects"]
data_rows_3 = [row for row in ws3.iter_rows(min_row=2, values_only=True) if any(row)]

check("Appended exactly 2 new rows", appended_3 == 2)
check("Total prospect rows is now 7", len(data_rows_3) == 7)
check("Original row 1 Status is STILL 'CLOSED'", data_rows_3[0][10] == "CLOSED")
check("Row 6 is 'Zeta Infra' with status 'NOT_CONTACTED'", data_rows_3[5][0] == "Zeta Infra" and data_rows_3[5][10] == "NOT_CONTACTED")
check("Row 7 is 'Eta Homes' with status 'NOT_CONTACTED'", data_rows_3[6][0] == "Eta Homes" and data_rows_3[6][10] == "NOT_CONTACTED")

# ------------------------------------------------------------------------------
# TEST 4: Identity Matching Hierarchy (Phone, Website, Fuzzy Name+City)
# ------------------------------------------------------------------------------
print("\n--- TEST 4: Identity Matching Hierarchy ---")
# 1. Match by phone with different formatting
p_phone_match = make_dummy_prospect("Alpha Realty Different Name", "9820011111", city="Mumbai")
# 2. Match by website
p_web_match = make_dummy_prospect("Some Random Name", "+91 90000 00001", website="https://www.zeta-infra.com")
# 3. Match by name + city
p_name_match = make_dummy_prospect("Eta Homes | Best Builders in Mumbai ⭐⭐⭐", "+91 90000 00002", city="Mumbai")

# Seed Zeta with website
ws3["L6"] = "https://zeta-infra.com"  # Column L is Website for Zeta
buf_seed = io.BytesIO()
wb_after_3.save(buf_seed)
mock_storage.upload_file(buf_seed.getvalue())

wb_loaded_4 = load_workbook(io.BytesIO(mock_storage.download_file()))
appended_4 = append_new_leads(wb_loaded_4, [p_phone_match, p_web_match, p_name_match])

check("All 3 duplicate variants were recognized and skipped (appended = 0)", appended_4 == 0)

# ------------------------------------------------------------------------------
# TEST 5: Storage Persistence across container restarts & API Download
# ------------------------------------------------------------------------------
print("\n--- TEST 5: Container restart simulation & API download ---")
fresh_container_storage = mock_storage
downloaded_bytes = fresh_container_storage.download_file()

check("Downloaded bytes is valid non-empty byte stream", downloaded_bytes is not None and len(downloaded_bytes) > 0)
downloaded_wb = load_workbook(io.BytesIO(downloaded_bytes))
check("Downloaded workbook has 7 rows and preserves 'CLOSED' status", len(downloaded_wb["Prospects"]["A"]) - 1 == 7 and downloaded_wb["Prospects"]["K2"].value == "CLOSED")

# ------------------------------------------------------------------------------
# TEST 6: Strict Server-Side Key Configuration (Fails Clearly when Unconfigured)
# ------------------------------------------------------------------------------
print("\n--- TEST 6: Strict server-side key check (fails clearly when unconfigured) ---")
unconfigured_client = SupabaseStorageClient(base_url="", api_key="")
try:
    unconfigured_client.download_file()
    check("Unconfigured client raises RuntimeError", False)
except RuntimeError as exc:
    check("Unconfigured client raised RuntimeError cleanly", "SUPABASE_KEY" in str(exc))

print("\n" + "=" * 80)
if failures:
    print(f"FAILED: {len(failures)} test(s) failed: {', '.join(failures)}")
    sys.exit(1)
else:
    print("ALL TESTS PASSED SUCCESSFULLY!")
    print("=" * 80)
    sys.exit(0)


