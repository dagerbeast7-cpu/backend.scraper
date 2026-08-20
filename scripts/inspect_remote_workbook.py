"""Download canonical workbook from live Railway API / Supabase Storage and inspect."""
import httpx
from openpyxl import load_workbook
import io
import phonenumbers

RAILWAY_URL = "https://backendscraper-production.up.railway.app/prospects/export/download"

try:
    print(f"Fetching from {RAILWAY_URL}...")
    res = httpx.get(RAILWAY_URL, timeout=30.0)
    print(f"HTTP Status: {res.status_code}")
    print(f"Content Length: {len(res.content)} bytes")
    
    if res.status_code == 200:
        wb = load_workbook(io.BytesIO(res.content))
        print("Sheets in workbook:", wb.sheetnames)
        ws = wb["Prospects"] if "Prospects" in wb.sheetnames else wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        print(f"Total rows in sheet (including header): {len(rows)}")
        header = rows[0]
        data_rows = rows[1:]
        print(f"Total prospect data rows: {len(data_rows)}")
        
        no_phone_count = 0
        valid_phone_count = 0
        
        for idx, row in enumerate(data_rows, start=2):
            bname = row[0]
            phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
            
            is_valid = False
            if phone:
                try:
                    parsed = phonenumbers.parse(phone, "IN")
                    if phonenumbers.is_valid_number(parsed):
                        is_valid = True
                except Exception:
                    is_valid = False
            
            if is_valid:
                valid_phone_count += 1
            else:
                no_phone_count += 1
                if no_phone_count <= 10:
                    print(f"  Row {idx}: {bname} | Phone: {phone}")
        
        print(f"\nSummary:")
        print(f"  Total Data Rows: {len(data_rows)}")
        print(f"  Rows with Usable Phone: {valid_phone_count}")
        print(f"  Rows with NO Usable Phone: {no_phone_count}")
        
except Exception as exc:
    print(f"Error fetching from Railway: {exc}")
