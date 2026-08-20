from __future__ import annotations

import io
import logging
import os
from datetime import datetime, timezone
from typing import Iterable, Sequence

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select

from app.config import settings
from app.db.base import get_session
from app.db.models import Prospect
from app.dedup.engine import (
    clean_business_name,
    normalize_name,
    normalize_phone,
    normalize_website,
)

logger = logging.getLogger(__name__)

EXCEL_COLUMNS = [
    ("Business Name", "business_name"),
    ("Contact / Broker Name", "contact_name"),
    ("Phone", "phone"),
    ("WhatsApp", "whatsapp"),
    ("Email", "email"),
    ("City", "city"),
    ("Area / Locality", "locality"),
    ("Address", "address"),
    ("Industry", "industry"),
    ("Score", "score"),
    ("Status", "status"),
    ("Website", "website"),
    ("Google Rating", "google_rating"),
    ("Reviews", "review_count"),
    ("WhatsApp Source", "whatsapp_source"),
    ("Description", "business_description"),
    ("Google Maps ID", "google_maps_id"),
]

STATUS_OPTIONS = ["NOT_CONTACTED", "MAYBE", "CONVERTED", "LOST"]
STATUS_MAP = {
    "NEW": "NOT_CONTACTED",
    "ENRICHING": "NOT_CONTACTED",
    "READY": "NOT_CONTACTED",
    "CONTACTED": "NOT_CONTACTED",
    "RESPONDED": "MAYBE",
    "INTERESTED": "MAYBE",
    "DEMO_BOOKED": "MAYBE",
    "CUSTOMER": "CONVERTED",
    "CONVERTED": "CONVERTED",
    "LOST": "LOST",
    "NOT_CONTACTED": "NOT_CONTACTED",
    "MAYBE": "MAYBE",
}

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
DATA_FONT = Font(size=10)


class SupabaseStorageClient:
    """
    Client for Supabase Storage REST API.
    Streams files in-memory to/from the canonical storage bucket.
    """

    def __init__(
        self,
        base_url: str | None = None,
        api_key: str | None = None,
        bucket_name: str | None = None,
    ):
        self.base_url = (base_url or settings.effective_supabase_url).rstrip("/")
        self.api_key = api_key or settings.effective_supabase_key
        self.bucket_name = bucket_name or settings.supabase_storage_bucket

    @property
    def is_configured(self) -> bool:
        return bool(self.base_url and self.api_key)

    def _headers(self, extra: dict | None = None) -> dict:
        h = {
            "Authorization": f"Bearer {self.api_key}",
            "apikey": self.api_key,
        }
        if extra:
            h.update(extra)
        return h

    def _require_configured(self) -> None:
        if not self.is_configured:
            missing = []
            if not self.base_url:
                missing.append("SUPABASE_URL (or DATABASE_URL with Supabase project host)")
            if not self.api_key:
                missing.append("SUPABASE_KEY (or SUPABASE_SERVICE_ROLE_KEY)")
            err_msg = f"Supabase Storage is not configured. Missing: {', '.join(missing)}"
            logger.error(err_msg)
            raise RuntimeError(err_msg)

    def ensure_bucket_exists(self) -> None:
        """Create bucket if it does not already exist."""
        self._require_configured()
        url = f"{self.base_url}/storage/v1/bucket"
        try:
            with httpx.Client(timeout=10.0) as client:
                res = client.post(
                    url,
                    headers=self._headers({"Content-Type": "application/json"}),
                    json={"id": self.bucket_name, "name": self.bucket_name, "public": False},
                )
                if res.status_code in (200, 201):
                    logger.info("Created Supabase Storage bucket %r", self.bucket_name)
                elif res.status_code in (400, 409):
                    logger.debug("Supabase Storage bucket %r already exists (HTTP %d)", self.bucket_name, res.status_code)
                else:
                    logger.warning(
                        "Supabase Storage bucket check returned HTTP %d: %s",
                        res.status_code,
                        res.text[:200],
                    )
        except Exception as exc:
            logger.warning("ensure_bucket_exists exception: %s", exc)

    def download_file(self, object_name: str | None = None) -> bytes | None:
        """
        Download file bytes from storage bucket.
        Returns None if object does not exist (404/not found).
        Raises RuntimeError on authentication/server errors.
        """
        self._require_configured()
        name = object_name or settings.supabase_storage_object
        url = f"{self.base_url}/storage/v1/object/{self.bucket_name}/{name}"

        try:
            with httpx.Client(timeout=30.0) as client:
                res = client.get(url, headers=self._headers())
                if res.status_code == 200:
                    logger.info("Downloaded %s (%d bytes) from Supabase Storage", name, len(res.content))
                    return res.content
                if res.status_code in (400, 404):
                    # Check if body indicates absent object vs genuine error
                    body_text = res.text.lower()
                    if "not found" in body_text or "not_found" in body_text or res.status_code == 404:
                        logger.info("Object %s not found in bucket %s (will be created)", name, self.bucket_name)
                        return None
                err = f"Supabase Storage download failed with HTTP {res.status_code}: {res.text[:300]}"
                logger.error(err)
                raise RuntimeError(err)
        except httpx.RequestError as exc:
            err = f"Supabase Storage network error during download: {exc}"
            logger.error(err)
            raise RuntimeError(err) from exc

    def upload_file(self, data: bytes, object_name: str | None = None) -> bool:
        """
        Upload file bytes to storage bucket atomically using x-upsert.
        Raises RuntimeError on authentication/server errors.
        """
        self._require_configured()
        name = object_name or settings.supabase_storage_object
        self.ensure_bucket_exists()

        url = f"{self.base_url}/storage/v1/object/{self.bucket_name}/{name}"
        headers = self._headers({
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "x-upsert": "true",
        })
        try:
            with httpx.Client(timeout=45.0) as client:
                res = client.post(url, headers=headers, content=data)
                if res.status_code in (200, 201):
                    logger.info("Successfully uploaded %s (%d bytes) to Supabase Storage", name, len(data))
                    return True
                # If POST rejected x-upsert with 400 on older versions, retry with PUT
                if res.status_code == 400:
                    put_res = client.put(url, headers=headers, content=data)
                    if put_res.status_code in (200, 201):
                        logger.info("Successfully updated %s (%d bytes) via PUT", name, len(data))
                        return True
                err = f"Supabase Storage upload failed with HTTP {res.status_code}: {res.text[:300]}"
                logger.error(err)
                raise RuntimeError(err)
        except httpx.RequestError as exc:
            err = f"Supabase Storage network error during upload: {exc}"
            logger.error(err)
            raise RuntimeError(err) from exc


def _build_identity_keys(prospect: Prospect) -> tuple[str | None, str | None, str | None, str | None]:
    """
    Extract the 4 normalized identity keys for a prospect:
    1. Google Maps ID
    2. Normalized phone
    3. Normalized website
    4. Normalized business name + city
    """
    g_id = str(prospect.google_maps_id).strip() if prospect.google_maps_id else None
    phone = normalize_phone(prospect.phone) if prospect.phone else None
    website = normalize_website(prospect.website) if prospect.website else None

    clean_name = clean_business_name(prospect.business_name) if prospect.business_name else None
    name_norm = normalize_name(clean_name) if clean_name else None
    city_norm = normalize_name(prospect.city) if prospect.city else ""
    name_city = f"{name_norm}::{city_norm}" if name_norm else None

    return g_id, phone, website, name_city


def _extract_existing_identities(ws) -> tuple[set[str], set[str], set[str], set[str]]:
    """
    Scan existing rows in the active Prospects sheet and build in-memory lookup sets.
    """
    existing_g_ids: set[str] = set()
    existing_phones: set[str] = set()
    existing_websites: set[str] = set()
    existing_name_cities: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        # Column indices (0-based):
        # 0: Business Name, 2: Phone, 5: City, 11: Website, 16: Google Maps ID
        bus_name = str(row[0]).strip() if len(row) > 0 and row[0] else None
        raw_phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
        city = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        raw_web = str(row[11]).strip() if len(row) > 11 and row[11] else None
        g_id = str(row[16]).strip() if len(row) > 16 and row[16] else None

        if g_id:
            existing_g_ids.add(g_id)
        if raw_phone:
            p_norm = normalize_phone(raw_phone)
            if p_norm:
                existing_phones.add(p_norm)
        if raw_web:
            w_norm = normalize_website(raw_web)
            if w_norm:
                existing_websites.add(w_norm)
        if bus_name:
            c_name = clean_business_name(bus_name)
            if c_name:
                n_norm = normalize_name(c_name)
                c_norm = normalize_name(city)
                if n_norm:
                    existing_name_cities.add(f"{n_norm}::{c_norm}")

    return existing_g_ids, existing_phones, existing_websites, existing_name_cities


def _apply_validation(ws, max_row: int) -> None:
    """Attach the Status dropdown DataValidation to Column K."""
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True,
    )
    dv.error = f"Please select a valid status: {', '.join(STATUS_OPTIONS)}"
    dv.errorTitle = "Invalid Status"
    dv.prompt = "Choose calling status"
    dv.promptTitle = "Status Options"
    ws.add_data_validation(dv)
    end_row = max(max_row + 500, 500)
    dv.add(f"K2:K{end_row}")


def _update_meta_sheet(wb: Workbook, total_prospects: int) -> None:
    """Create or update the 'Export Info' metadata sheet."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    if "Export Info" in wb.sheetnames:
        meta = wb["Export Info"]
        meta.delete_rows(1, meta.max_row)
    else:
        meta = wb.create_sheet("Export Info")

    meta.append(["Last Updated", now])
    meta.append(["Total Prospects", total_prospects])
    meta.append(["Status Options", ", ".join(STATUS_OPTIONS)])
    meta.append(["Storage Location", f"{settings.supabase_storage_bucket}/{settings.supabase_storage_object}"])


def list_no_phone_rows(wb: Workbook) -> list[dict]:
    """
    Identify existing rows in the Prospects sheet that have no usable phone number.
    Returns list of dicts with row number, business name, city, and raw phone.
    """
    ws = wb["Prospects"] if "Prospects" in wb.sheetnames else wb.active
    no_phone_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        bus_name = str(row[0]).strip() if len(row) > 0 and row[0] else ""
        raw_phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
        city = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        if not raw_phone or normalize_phone(raw_phone) is None:
            no_phone_rows.append({
                "row_number": row_idx,
                "business_name": bus_name,
                "raw_phone": raw_phone,
                "city": city,
            })
    return no_phone_rows


def create_initial_workbook(prospects: Sequence[Prospect]) -> Workbook:
    """Create the initial formatted canonical workbook containing ONLY leads with usable phone numbers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    # Filter strictly to prospects with a usable phone number and business name
    eligible_prospects = [
        p for p in prospects
        if p.business_name and normalize_phone(p.phone) is not None
    ]

    # 1. Write Headers
    for col_idx, (header, _) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # 2. Write Data Rows
    for row_idx, prospect in enumerate(eligible_prospects, 2):
        normalized_phone = normalize_phone(prospect.phone)
        for col_idx, (_, attr) in enumerate(EXCEL_COLUMNS, 1):
            val = getattr(prospect, attr, None)
            if hasattr(val, "value"):
                val = val.value
            if attr == "status":
                val = STATUS_MAP.get(str(val).upper(), "NOT_CONTACTED")
            elif attr == "phone":
                val = normalized_phone
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT

    max_row = max(len(eligible_prospects) + 1, 1)
    _apply_validation(ws, max_row)

    # 3. Auto-fit column widths
    for col_idx, (header, _) in enumerate(EXCEL_COLUMNS, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

    # Freeze header row & auto-filter
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Metadata sheet
    _update_meta_sheet(wb, len(eligible_prospects))
    return wb


def append_new_leads(wb: Workbook, candidates: Iterable[Prospect]) -> int:
    """
    Append only genuinely new prospects with USABLE PHONE NUMBERS to the existing workbook.
    NEVER overwrites existing rows or caller edits.
    Skips candidates without a usable phone number.
    Returns count of new rows appended.
    """
    ws = wb["Prospects"] if "Prospects" in wb.sheetnames else wb.active
    existing_g_ids, existing_phones, existing_websites, existing_name_cities = _extract_existing_identities(ws)

    appended_count = 0
    current_row = ws.max_row

    for prospect in candidates:
        # Rule: Only prospects with a usable phone number are eligible for Excel
        if not prospect.business_name:
            continue

        normalized_phone = normalize_phone(prospect.phone)
        if not normalized_phone:
            continue

        g_id, _, website, name_city = _build_identity_keys(prospect)

        # 4-tier match strategy
        if g_id and g_id in existing_g_ids:
            continue
        if normalized_phone in existing_phones:
            continue
        if website and website in existing_websites:
            continue
        if name_city and name_city in existing_name_cities:
            continue

        # Genuinely new with usable phone -> Append row
        current_row += 1
        for col_idx, (_, attr) in enumerate(EXCEL_COLUMNS, 1):
            val = getattr(prospect, attr, None)
            if hasattr(val, "value"):
                val = val.value
            if attr == "status":
                val = "NOT_CONTACTED"  # Initial caller status for newly appended leads
            elif attr == "phone":
                val = normalized_phone
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = DATA_FONT

        # Add new keys to lookup sets so intra-batch duplicates are skipped
        if g_id:
            existing_g_ids.add(g_id)
        existing_phones.add(normalized_phone)
        if website:
            existing_websites.add(website)
        if name_city:
            existing_name_cities.add(name_city)

        appended_count += 1

    if appended_count > 0:
        _apply_validation(ws, current_row)
        ws.auto_filter.ref = f"A1:{ws.cell(row=current_row, column=len(EXCEL_COLUMNS)).coordinate}"

    total_leads = max(current_row - 1, 0)
    _update_meta_sheet(wb, total_leads)
    return appended_count


def sync_prospects_to_storage_workbook(
    candidate_prospects: Sequence[Prospect] | None = None,
    storage_client: SupabaseStorageClient | None = None,
) -> dict:
    """
    Core synchronization workflow:
    1. Download canonical workbook from Supabase Storage.
    2. If missing: create initial workbook from DB and upload.
    3. If present: preserve all existing rows, append only new leads, and upload back.
    """
    client = storage_client or SupabaseStorageClient()
    existing_bytes = client.download_file()

    with get_session() as session:
        if candidate_prospects is None or existing_bytes is None:
            # Query all prospects ordered by score if initial creation or no candidates passed
            prospects_to_check = session.execute(
                select(Prospect).order_by(Prospect.score.desc())
            ).scalars().all()
        else:
            prospects_to_check = candidate_prospects

        if existing_bytes is None:
            logger.info("No canonical workbook found in Supabase Storage. Creating initial workbook.")
            wb = create_initial_workbook(prospects_to_check)
            appended = len(prospects_to_check)
        else:
            wb = load_workbook(filename=io.BytesIO(existing_bytes))
            appended = append_new_leads(wb, prospects_to_check)

        out_buf = io.BytesIO()
        wb.save(out_buf)
        out_bytes = out_buf.getvalue()

        # Upload updated workbook back to Supabase Storage
        uploaded = client.upload_file(out_bytes)

    result = {
        "appended": appended,
        "total_prospects": len(wb["Prospects"]["A"]) - 1 if "Prospects" in wb.sheetnames else 0,
        "uploaded_to_storage": uploaded,
        "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
    }
    logger.info("sync_prospects_to_storage_workbook finished: %s", result)
    return result


def get_canonical_workbook_bytes(storage_client: SupabaseStorageClient | None = None) -> bytes:
    """
    Fetch the latest canonical workbook bytes from Supabase Storage.
    If absent, generates initial workbook and uploads it.
    Fails clearly if Supabase Storage cannot be accessed or updated.
    """
    client = storage_client or SupabaseStorageClient()
    data = client.download_file()
    if data is not None:
        return data

    # Absent -> Generate initial workbook, upload to Supabase Storage, and return
    res = sync_prospects_to_storage_workbook(storage_client=client)
    logger.info("Generated and uploaded initial workbook: %s", res)
    refreshed = client.download_file()
    if refreshed is not None:
        return refreshed

    raise RuntimeError("Failed to verify canonical workbook in Supabase Storage after upload.")

