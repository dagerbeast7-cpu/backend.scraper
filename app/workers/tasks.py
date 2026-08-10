from __future__ import annotations

import logging
from typing import Optional

from app.db.base import get_session
from app.dedup.engine import DedupEngine, is_mobile_number
from app.enrichment.service import EnrichmentService
from app.scoring.engine import apply_score
from app.scraper.base import PHASE_1_ICP_QUERIES
from app.scraper.factory import get_provider
from app.workers.celery_app import celery_app

logger = logging.getLogger(__name__)


@celery_app.task(name="app.workers.tasks.run_scrape_pipeline", bind=True, max_retries=2)
def run_scrape_pipeline(
    self,
    city: str,
    queries: Optional[list[str]] = None,
    max_results_per_query: int = 60,
    provider: Optional[str] = None,
    area: Optional[str] = None,
) -> dict:
    """
    Full pipeline for one city (optionally narrowed to one area/locality
    within it, e.g. "Bandra West" within "Mumbai"): scrape -> dedup ->
    enrich -> score -> save. Runs as a Celery task so it can be triggered
    from the API or on the nightly beat schedule without blocking a web
    request.
    """
    queries = queries or PHASE_1_ICP_QUERIES
    scraper = get_provider(provider)

    created_count = 0
    updated_count = 0
    total_seen = 0

    with get_session() as session:
        dedup = DedupEngine(session)
        enrichment = EnrichmentService()

        try:
            for query in queries:
                try:
                    leads = scraper.search(query, city, max_results_per_query, area=area)
                except Exception:  # noqa: BLE001
                    logger.exception(
                        "Scrape failed for query=%r city=%r area=%r", query, city, area
                    )
                    continue

                total_seen += len(leads)

                for lead in leads:
                    try:
                        prospect, created = dedup.upsert(lead)
                        if created:
                            created_count += 1
                        else:
                            updated_count += 1

                        try:
                            enrichment.enrich(prospect)
                        except Exception:  # noqa: BLE001
                            logger.exception("Enrichment failed for %s", prospect.business_name)

                        if not prospect.whatsapp and is_mobile_number(prospect.phone):
                            prospect.whatsapp = prospect.phone
                            prospect.whatsapp_source = "inferred"

                        apply_score(prospect)
                        session.commit()
                    except Exception:  # noqa: BLE001
                        session.rollback()
                        logger.exception("Failed to process lead %s, rolled back session", lead.business_name)
        finally:
            enrichment.close()

    result = {
        "city": city,
        "area": area,
        "queries": queries,
        "total_seen": total_seen,
        "created": created_count,
        "updated": updated_count,
    }
    logger.info("run_scrape_pipeline finished: %s", result)
    return result


@celery_app.task(name="app.workers.tasks.export_prospects_to_excel")
def export_prospects_to_excel() -> dict:
    """
    Export all prospects to an Excel file that cold callers can use.
    Runs on a schedule (every hour by default) via Celery Beat.
    The file is saved to /code/exports/leadzen_prospects.xlsx which
    maps to the host's ./exports/ directory via the Docker volume mount.
    """
    import os
    from datetime import datetime, timezone

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from sqlalchemy import select

    from app.db.base import get_session
    from app.db.models import Prospect

    export_dir = "/code/exports"
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, "leadzen_prospects.xlsx")

    columns = [
        ("Business Name", "business_name"),
        ("Contact / Broker Name", "contact_name"),
        ("Phone", "phone"),
        ("WhatsApp", "whatsapp"),
        ("Email", "email"),
        ("City", "city"),
        ("Area / Locality", "locality"),
        ("Address", "address"),
        ("Industry", "industry"),
        ("Score", "score"),
        ("Status", "status"),
        ("Website", "website"),
        ("Google Rating", "google_rating"),
        ("Reviews", "review_count"),
        ("WhatsApp Source", "whatsapp_source"),
        ("Description", "business_description"),
    ]

    with get_session() as session:
        prospects = session.execute(
            select(Prospect).order_by(Prospect.score.desc())
        ).scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Prospects"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

        # Status mapping for caller options: NOT_CONTACTED, MAYBE, CONVERTED, LOST
        STATUS_MAP = {
            "NEW": "NOT_CONTACTED",
            "ENRICHING": "NOT_CONTACTED",
            "READY": "NOT_CONTACTED",
            "CONTACTED": "NOT_CONTACTED",
            "RESPONDED": "MAYBE",
            "INTERESTED": "MAYBE",
            "DEMO_BOOKED": "MAYBE",
            "CUSTOMER": "CONVERTED",
            "CONVERTED": "CONVERTED",
            "LOST": "LOST",
            "NOT_CONTACTED": "NOT_CONTACTED",
            "MAYBE": "MAYBE",
        }

        # Write headers
        for col_idx, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, prospect in enumerate(prospects, 2):
            for col_idx, (_, attr) in enumerate(columns, 1):
                value = getattr(prospect, attr, None)
                # Convert enums to string & map status values
                if hasattr(value, "value"):
                    value = value.value
                if attr == "status":
                    value = STATUS_MAP.get(str(value).upper(), "NOT_CONTACTED")
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Add Data Validation (Dropdown menu in Excel for Status column K)
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type="list",
            formula1='"NOT_CONTACTED,MAYBE,CONVERTED,LOST"',
            allow_blank=True,
        )
        dv.error = "Please select a valid status: NOT_CONTACTED, MAYBE, CONVERTED, or LOST"
        dv.errorTitle = "Invalid Status"
        dv.prompt = "Choose calling status"
        dv.promptTitle = "Status Options"

        ws.add_data_validation(dv)
        max_row = max(len(prospects) + 1, 500)
        dv.add(f"K2:K{max_row}")

        # Auto-fit column widths (approximate)
        for col_idx, (header, _) in enumerate(columns, 1):
            max_len = len(header)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 50))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

        # Add metadata sheet with export timestamp
        meta = wb.create_sheet("Export Info")
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        meta.append(["Last Updated", now])
        meta.append(["Total Prospects", len(prospects)])
        meta.append(["Status Options", "NOT_CONTACTED, MAYBE, CONVERTED, LOST"])
        meta.append(["File", filepath])

        # Freeze top row and enable auto-filter on the data sheet
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        wb.save(filepath)

    logger.info(
        "Exported %d prospects to %s at %s",
        len(prospects), filepath, now,
    )
    return {"exported": len(prospects), "file": filepath, "timestamp": now}


@celery_app.task(name="app.workers.tasks.run_nightly_region_scrape")
def run_nightly_region_scrape(region_key: str) -> dict:
    """
    Automated nightly scheduled batch task. Scrapes all micro-market target
    localities defined for `region_key` ('delhi_ncr', 'mumbai', 'bangalore')
    during the 8:00 PM - 6:00 AM IST window.
    """
    from app.scraper.base import NIGHTLY_SCHEDULE_TARGETS

    targets = NIGHTLY_SCHEDULE_TARGETS.get(region_key, [])
    if not targets:
        logger.warning("No targets found for region_key=%r", region_key)
        return {"region": region_key, "processed": 0}

    total_created = 0
    total_updated = 0

    for target in targets:
        city = target["city"]
        area = target.get("area")
        try:
            res = run_scrape_pipeline(
                city=city,
                area=area,
                max_results_per_query=40,
            )
            total_created += res.get("created", 0)
            total_updated += res.get("updated", 0)
        except Exception:  # noqa: BLE001
            logger.exception("Nightly batch failed for city=%r area=%r", city, area)

    # Automatically refresh Excel sheet after the region batch completes
    export_prospects_to_excel()

    result = {
        "region": region_key,
        "targets_count": len(targets),
        "total_created": total_created,
        "total_updated": total_updated,
    }
    logger.info("run_nightly_region_scrape finished: %s", result)
    return result

