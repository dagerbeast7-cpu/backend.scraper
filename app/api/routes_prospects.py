from __future__ import annotations

from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.schemas import ProspectOut, ProspectStatusUpdate, ScrapeJobRequest, ScrapeJobResponse
from app.db.base import SessionLocal
from app.db.models import LeadStatus, Prospect
from app.scraper.base import PHASE_1_ICP_QUERIES

router = APIRouter(prefix="/prospects", tags=["prospects"])


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.get("", response_model=list[ProspectOut])
def list_prospects(
    city: Optional[str] = None,
    locality: Optional[str] = None,
    state: Optional[str] = None,
    industry: Optional[str] = None,
    status: Optional[str] = None,
    min_score: Optional[int] = Query(None, ge=0, le=100),
    has_email: Optional[bool] = None,
    has_whatsapp: Optional[bool] = None,
    has_website: Optional[bool] = None,
    limit: int = Query(50, le=500),
    offset: int = 0,
    db: Session = Depends(get_db),
):
    stmt = select(Prospect)

    if city:
        stmt = stmt.where(Prospect.city.ilike(city))
    if locality:
        stmt = stmt.where(Prospect.locality.ilike(f"%{locality}%"))
    if state:
        stmt = stmt.where(Prospect.state.ilike(state))
    if industry:
        stmt = stmt.where(Prospect.industry.ilike(f"%{industry}%"))
    if status:
        stmt = stmt.where(Prospect.status == status.upper())
    if min_score is not None:
        stmt = stmt.where(Prospect.score >= min_score)
    if has_email is not None:
        stmt = stmt.where(Prospect.email.isnot(None) if has_email else Prospect.email.is_(None))
    if has_whatsapp is not None:
        stmt = stmt.where(
            Prospect.whatsapp.isnot(None) if has_whatsapp else Prospect.whatsapp.is_(None)
        )
    if has_website is not None:
        stmt = stmt.where(
            Prospect.website.isnot(None) if has_website else Prospect.website.is_(None)
        )

    stmt = stmt.order_by(Prospect.score.desc()).offset(offset).limit(limit)
    return db.execute(stmt).scalars().all()


@router.get("/{prospect_id}", response_model=ProspectOut)
def get_prospect(prospect_id: UUID, db: Session = Depends(get_db)):
    prospect = db.get(Prospect, prospect_id)
    if not prospect:
        raise HTTPException(status_code=404, detail="Prospect not found")
    return prospect


@router.patch("/{prospect_id}/status", response_model=ProspectOut)
def update_status(prospect_id: UUID, payload: ProspectStatusUpdate, db: Session = Depends(get_db)):
    prospect = db.get(Prospect, prospect_id)
    if not prospect:
        raise HTTPException(status_code=404, detail="Prospect not found")
    try:
        prospect.status = LeadStatus(payload.status.upper())
    except ValueError:
        valid = ", ".join(s.value for s in LeadStatus)
        raise HTTPException(status_code=400, detail=f"Invalid status. Valid values: {valid}")
    db.commit()
    db.refresh(prospect)
    return prospect


@router.post("/scrape", response_model=ScrapeJobResponse, status_code=202)
def trigger_scrape(payload: ScrapeJobRequest):
    from app.workers.tasks import run_scrape_pipeline

    queries = payload.queries or PHASE_1_ICP_QUERIES
    async_result = run_scrape_pipeline.delay(
        city=payload.city,
        area=payload.area,
        queries=queries,
        max_results_per_query=payload.max_results_per_query,
        provider=payload.provider,
    )
    return ScrapeJobResponse(
        task_id=async_result.id, city=payload.city, area=payload.area, queries=queries
    )


@router.post("/export", status_code=202)
def trigger_export():
    """Trigger an immediate Excel export (runs in the background)."""
    from app.workers.tasks import export_prospects_to_excel

    async_result = export_prospects_to_excel.delay()
    return {"task_id": async_result.id, "message": "Excel export started"}


@router.get("/export/download")
def download_export():
    """Download the latest canonical Excel file directly from Supabase Storage."""
    from fastapi.responses import Response
    from app.storage.excel_storage import get_canonical_workbook_bytes

    try:
        data = get_canonical_workbook_bytes()
    except Exception as err:
        raise HTTPException(status_code=500, detail=f"Failed to fetch canonical Excel export: {err}")

    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="leadzen_prospects.xlsx"'},
    )


@router.post("/import-excel")
def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Import an edited Excel sheet from a cold caller to update prospect statuses
    and contact details in the database.
    """
    import io
    from openpyxl import load_workbook
    from app.dedup.engine import normalize_phone

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel spreadsheet (.xlsx)")

    contents = file.file.read()
    wb = load_workbook(filename=io.BytesIO(contents))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers) if h}

    phone_idx = header_map.get("phone")
    name_idx = header_map.get("business name")
    status_idx = header_map.get("status")
    contact_idx = header_map.get("contact / broker name")

    if status_idx is None:
        raise HTTPException(status_code=400, detail="Excel sheet must contain a 'Status' column")

    updated_count = 0
    valid_statuses = {"NOT_CONTACTED", "MAYBE", "CONVERTED", "LOST", "NEW", "CUSTOMER", "INTERESTED"}

    for row in ws.iter_rows(min_row=2, values_only=True):
        status_val = str(row[status_idx]).strip().upper() if row[status_idx] else None
        if not status_val or status_val not in valid_statuses:
            continue

        raw_phone = str(row[phone_idx]).strip() if phone_idx is not None and row[phone_idx] else None
        phone = normalize_phone(raw_phone)
        bus_name = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else None

        prospect = None
        if phone:
            prospect = db.execute(select(Prospect).where(Prospect.phone == phone)).scalar_one_or_none()
        if not prospect and bus_name:
            prospect = db.execute(select(Prospect).where(Prospect.business_name.ilike(bus_name))).scalar_one_or_none()

        if prospect:
            try:
                prospect.status = LeadStatus(status_val)
            except ValueError:
                # Fallback mapping
                if status_val == "NEW":
                    prospect.status = LeadStatus.NOT_CONTACTED
                elif status_val in ("INTERESTED", "MAYBE"):
                    prospect.status = LeadStatus.MAYBE
                elif status_val in ("CUSTOMER", "CONVERTED"):
                    prospect.status = LeadStatus.CONVERTED

            if contact_idx is not None and row[contact_idx]:
                prospect.contact_name = str(row[contact_idx]).strip()

            updated_count += 1

    db.commit()
    return {"updated": updated_count, "message": f"Successfully updated {updated_count} prospect statuses from Excel"}
